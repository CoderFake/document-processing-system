import os
import json
import tempfile
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import uuid
import asyncpg
import logging

from domain.models import ExcelDocumentInfo, ExcelTemplateInfo, BatchProcessingInfo, MergeInfo
from domain.exceptions import DocumentNotFoundException, TemplateNotFoundException, StorageException
from infrastructure.minio_client import MinioClient
from core.config import settings

logger = logging.getLogger(__name__)

DOCUMENTS_TABLE = "documents"

class ExcelDocumentRepository:
    """
    Repository để làm việc với tài liệu Excel, lưu trữ doc_metadata trong PostgreSQL.
    """

    def __init__(self, pool: asyncpg.Pool):
        """
        Khởi tạo repository.

        Args:
            pool: Connection pool của AsyncPG
        """
        self.pool = pool

    async def _get_sheet_info(self, content: bytes) -> Tuple[List[str], int]:
        """
        Đọc tên các sheet và số lượng sheet từ nội dung tài liệu Excel.

        Args:
            content: Nội dung tài liệu Excel

        Returns:
            Tuple (danh sách tên sheet, số lượng sheet)
        """
        try:
            suffix = ".xlsx"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
                temp.write(content)
                temp_path = temp.name

            try:
                import openpyxl
                wb = openpyxl.load_workbook(temp_path, read_only=True)
                sheet_names = wb.sheetnames
                sheet_count = len(sheet_names)
                wb.close()
                return sheet_names, sheet_count
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
        except Exception as e:
           
            return [], 0

    async def _serialize_metadata(self, doc_metadata: Optional[Dict[str, Any]]) -> Optional[str]:
        if doc_metadata is None:
            return None
        return json.dumps(doc_metadata)

    async def _deserialize_metadata(self, doc_metadata_str: Optional[str]) -> Optional[Dict[str, Any]]:
        if doc_metadata_str is None:
            return None
        try:
            return json.loads(doc_metadata_str)
        except json.JSONDecodeError:
            logger.warning(f"Could not deserialize doc_metadata: {doc_metadata_str}")
            return None

    async def save(self, doc_info: ExcelDocumentInfo) -> ExcelDocumentInfo:
        async with self.pool.acquire() as connection:
            doc_info.id = doc_info.id or str(uuid.uuid4())
            doc_info.storage_id = doc_info.storage_id or str(uuid.uuid4())
            doc_info.created_at = doc_info.created_at or datetime.utcnow()
            doc_info.updated_at = doc_info.updated_at or datetime.utcnow()
            doc_info.version = doc_info.version or 1
            doc_info.document_category = "excel"

            serialized_doc_metadata = await self._serialize_metadata(doc_info.doc_metadata)

            query = f"""
                INSERT INTO {DOCUMENTS_TABLE} (
                    id, storage_id, document_category, title, description, 
                    file_size, file_type, storage_path, original_filename, doc_metadata, 
                    created_at, updated_at, user_id, version, checksum, sheet_count
                )
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16)
                ON CONFLICT (id) DO UPDATE SET
                    title = EXCLUDED.title,
                    description = EXCLUDED.description,
                    file_size = EXCLUDED.file_size,
                    file_type = EXCLUDED.file_type,
                    storage_path = EXCLUDED.storage_path,
                    original_filename = EXCLUDED.original_filename,
                    doc_metadata = EXCLUDED.doc_metadata,
                    updated_at = EXCLUDED.updated_at,
                    user_id = EXCLUDED.user_id,
                    version = EXCLUDED.version,
                    checksum = EXCLUDED.checksum,
                    sheet_count = EXCLUDED.sheet_count
                RETURNING *;
        """
        try:
            row = await connection.fetchrow(
                query,
                doc_info.id,
                doc_info.storage_id,
                doc_info.document_category,
                doc_info.title,
                doc_info.description,
                doc_info.file_size,
                doc_info.file_type,
                doc_info.storage_path,
                doc_info.original_filename,
                serialized_doc_metadata,
                doc_info.created_at,
                doc_info.updated_at,
                doc_info.user_id,
                doc_info.version,
                doc_info.checksum,
                doc_info.sheet_count
            )
            if row:
                saved_doc_info = ExcelDocumentInfo(**{k: v for k, v in row.items() if k != 'doc_metadata'})
                saved_doc_info.doc_metadata = await self._deserialize_metadata(row.get('doc_metadata'))
                logger.info(f"Saved/Updated document {saved_doc_info.id} for user {saved_doc_info.user_id}")
                return saved_doc_info
            else:
                logger.error(f"Failed to save/update document {doc_info.id}, no row returned.")
                raise StorageException(f"Failed to save/update document {doc_info.id}")
        except asyncpg.UniqueViolationError as e:
            logger.error(f"Unique constraint violation for doc {doc_info.id}: {e}")
            raise StorageException(f"Error saving document due to unique constraint: {e}")
        except Exception as e:
            logger.error(f"Error saving/updating document {doc_info.id}: {e}", exc_info=True)
            raise StorageException(f"Could not save/update document {doc_info.id}: {e}")

    async def get_by_id(self, doc_id: str, user_id: str) -> Optional[ExcelDocumentInfo]:
        async with self.pool.acquire() as connection:
            query = f"SELECT * FROM {DOCUMENTS_TABLE} WHERE id = $1 AND user_id = $2 AND document_category = 'excel';"
            row = await connection.fetchrow(query, doc_id, user_id)
            if row:
                doc_info = ExcelDocumentInfo(**{k: v for k, v in row.items() if k != 'doc_metadata'})
                doc_info.doc_metadata = await self._deserialize_metadata(row.get('doc_metadata'))
                return doc_info
            return None

    async def list_by_user_id(
        self, user_id: str, skip: int = 0, limit: int = 20, 
        search_term: Optional[str] = None, sort_by: str = 'created_at', sort_order: str = 'desc'
    ) -> Tuple[List[ExcelDocumentInfo], int]:
        async with self.pool.acquire() as connection:
            base_query = f"FROM {DOCUMENTS_TABLE} WHERE user_id = $1 AND document_category = 'excel'"
            params = [user_id]
            param_idx = 2

            if search_term:
                base_query += f" AND (title ILIKE ${param_idx} OR original_filename ILIKE ${param_idx})"
                params.append(f"%{search_term}%")
                param_idx += 1

            count_query = f"SELECT COUNT(*) {base_query};"
            total_count_row = await connection.fetchrow(count_query, *params)
            total_count = total_count_row['count'] if total_count_row else 0


            allowed_sort_fields = ['title', 'created_at', 'updated_at', 'file_size', 'original_filename']
            if sort_by not in allowed_sort_fields:
                sort_by = 'created_at'
            if sort_order.lower() not in ['asc', 'desc']:
                sort_order = 'desc'

            data_query = f"""
                SELECT * {base_query}
                ORDER BY {sort_by} {sort_order.upper()}
                LIMIT ${param_idx} OFFSET ${param_idx + 1};
            """
            params.extend([limit, skip])
            
            rows = await connection.fetch(data_query, *params)
            documents = []
            for row in rows:
                doc_info = ExcelDocumentInfo(**{k: v for k, v in row.items() if k != 'doc_metadata'})
                doc_info.doc_metadata = await self._deserialize_metadata(row.get('doc_metadata'))
                documents.append(doc_info)
            return documents, total_count

    async def update_metadata(self, doc_id: str, user_id: str, update_data: Dict[str, Any]) -> Optional[ExcelDocumentInfo]:
        async with self.pool.acquire() as connection:
            allowed_fields = {'title', 'description', 'doc_metadata', 'sheet_count', 'original_filename'}
            set_clauses = []
            values = []
            param_idx = 1

            for key, value in update_data.items():
                if key in allowed_fields:
                    if key == 'doc_metadata':
                        values.append(await self._serialize_metadata(value))
                    else:
                        values.append(value)
                    set_clauses.append(f"{key} = ${param_idx}")
                    param_idx += 1
            
            if not set_clauses:
                logger.warning(f"Update doc_metadata for {doc_id} called with no valid fields.")
                return await self.get_by_id(doc_id, user_id)

            set_clauses.append(f"updated_at = ${param_idx}")
            values.append(datetime.utcnow())
            param_idx += 1

            values.extend([doc_id, user_id])

            query = f"""
                UPDATE {DOCUMENTS_TABLE}
                SET {', '.join(set_clauses)}
                WHERE id = ${param_idx} AND user_id = ${param_idx + 1} AND document_category = 'excel'
                RETURNING *;
            """
            
            updated_row = await connection.fetchrow(query, *values)
            if updated_row:
                doc_info = ExcelDocumentInfo(**{k: v for k, v in updated_row.items() if k != 'doc_metadata'})
                doc_info.doc_metadata = await self._deserialize_metadata(updated_row.get('doc_metadata'))
                logger.info(f"Updated doc_metadata for document {doc_id} for user {user_id}")
                return doc_info
            logger.warning(f"Document {doc_id} not found for user {user_id} for doc_metadata update.")
            return None

    async def delete(self, doc_id: str, user_id: str) -> bool:
        async with self.pool.acquire() as connection:


            query = f"DELETE FROM {DOCUMENTS_TABLE} WHERE id = $1 AND user_id = $2 AND document_category = 'excel' RETURNING id;"
            deleted_id = await connection.fetchval(query, doc_id, user_id)
            
            if deleted_id:
                logger.info(f"Deleted document {doc_id} for user {user_id} from DB.")
                return True 
            logger.warning(f"Document {doc_id} not found for user {user_id} for deletion or not an excel document.")
            return False

    async def check_exists(self, doc_id: str, user_id: str) -> bool:
        async with self.pool.acquire() as connection:
            query = f"SELECT EXISTS(SELECT 1 FROM {DOCUMENTS_TABLE} WHERE id = $1 AND user_id = $2 AND document_category = 'excel');"
            exists = await connection.fetchval(query, doc_id, user_id)
            return bool(exists)

class ExcelTemplateRepository:
    """
    Repository để làm việc với mẫu tài liệu Excel.
    """

    def __init__(self, minio_client: MinioClient):
        """
        Khởi tạo repository.

        Args:
            minio_client: Client MinIO để lưu trữ mẫu
        """
        self.minio_client = minio_client
        self.templates_metadata_file = os.path.join(settings.TEMP_DIR, "excel_templates_metadata.json")
        self.templates: Dict[str, ExcelTemplateInfo] = {}
        self._load_metadata()

    def _load_metadata(self) -> None:
        """
        Tải doc_metadata của mẫu từ file.
        """
        try:
            if os.path.exists(self.templates_metadata_file):
                with open(self.templates_metadata_file, "r") as f:
                    data = json.load(f)
                    for template_id, template_data in data.items():
                        self.templates[template_id] = ExcelTemplateInfo(**template_data)
        except Exception as e:
            print(f"Error loading Excel template doc_metadata: {e}, creating new file if it doesn't exist.")
            if not os.path.exists(self.templates_metadata_file):
                self._save_metadata()

    def _save_metadata(self) -> None:
        """
        Lưu doc_metadata của mẫu vào file.
        """
        try:
            os.makedirs(settings.TEMP_DIR, exist_ok=True)
            data = {template_id: template.dict() for template_id, template in self.templates.items() if hasattr(template, 'dict')}
            if not data and self.templates:
                 data = {
                    tid: {
                        "id": t.id, "name": t.name, "description": t.description, 
                        "file_size": t.file_size, "category": t.category,
                        "storage_path": t.storage_path, "created_at": t.created_at.isoformat() if t.created_at else None,
                        "updated_at": t.updated_at.isoformat() if t.updated_at else None, 
                        "variables": t.variables, "sample_data": t.sample_data
                    } for tid, t in self.templates.items()
                }

            with open(self.templates_metadata_file, "w") as f:
                json.dump(data, f, default=str)
        except Exception as e:
            raise StorageException(f"Không thể lưu doc_metadata mẫu Excel: {str(e)}")

    async def save(self, template_info: ExcelTemplateInfo, content: bytes) -> ExcelTemplateInfo:
        """
        Lưu mẫu mới.
        """
        try:
           
            minio_object_name = f"templates/{template_info.category}/{template_info.id}/{template_info.name}.xlsx"
            
            await self.minio_client.upload_template(
                content=content,
                
                object_name_override=minio_object_name
            )

            template_info.storage_path = minio_object_name
            template_info.file_size = len(content)

            self.templates[template_info.id] = template_info
            self._save_metadata()
            return template_info
        except Exception as e:
            raise StorageException(f"Không thể lưu mẫu Excel: {str(e)}")

    async def _get_sheet_names(self, content: bytes) -> List[str]:
        """Helper to get sheet names, similar to ExcelDocumentRepository"""
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp:
                temp.write(content)
                temp_path = temp.name
            try:
                import openpyxl
                wb = openpyxl.load_workbook(temp_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                return sheet_names
            finally:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
        except Exception:
            return []

    async def get(self, template_id: str) -> Tuple[ExcelTemplateInfo, bytes]:
        """
        Lấy thông tin và nội dung mẫu.
        """
        try:
            if template_id not in self.templates:
                raise TemplateNotFoundException(template_id)
            template_info = self.templates[template_id]
            content = await self.minio_client.download_template(template_info.storage_path) # Giả sử có hàm riêng
            return template_info, content
        except TemplateNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể lấy mẫu Excel {template_id}: {str(e)}")

    async def update(self, template_info: ExcelTemplateInfo, content: Optional[bytes] = None) -> ExcelTemplateInfo:
        """
        Cập nhật thông tin mẫu. Nếu content được cung cấp, upload lại.
        """
        try:
            if template_info.id not in self.templates:
                raise TemplateNotFoundException(template_info.id)

            existing_template = self.templates[template_info.id]
            template_info.updated_at = datetime.now()

            if content:
                minio_object_name = f"templates/{template_info.category}/{template_info.id}/{template_info.name}.xlsx"
                await self.minio_client.upload_template(
                    content=content,
                    object_name_override=minio_object_name
                )
                template_info.storage_path = minio_object_name
                template_info.file_size = len(content)
               
            else:
                template_info.storage_path = existing_template.storage_path
                template_info.file_size = existing_template.file_size

            self.templates[template_info.id] = template_info
            self._save_metadata()
            return template_info
        except TemplateNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể cập nhật mẫu Excel {template_info.id}: {str(e)}")

    async def delete(self, template_id: str) -> None:
        """
        Xóa mẫu.
        """
        try:
            if template_id not in self.templates:
                raise TemplateNotFoundException(template_id)
            template_info = self.templates[template_id]
            await self.minio_client.delete_template(template_info.storage_path) # Giả sử có hàm riêng
            del self.templates[template_id]
            self._save_metadata()
        except TemplateNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể xóa mẫu Excel {template_id}: {str(e)}")

    async def list(self, category: Optional[str] = None, skip: int = 0, limit: int = 10) -> List[ExcelTemplateInfo]:
        """
        Lấy danh sách mẫu.
        """
        try:
            filtered_templates = []
            for template in self.templates.values():
                if category:
                    if template.category.lower() == category.lower():
                        filtered_templates.append(template)
                else:
                    filtered_templates.append(template)

            sorted_templates = sorted(
                filtered_templates,
                key=lambda x: (x.name.lower() if x.name else '', x.created_at),
                reverse=False 
            )
            return sorted_templates[skip:skip + limit]
        except Exception as e:
            raise StorageException(f"Không thể lấy danh sách mẫu Excel: {str(e)}")


class BatchProcessingRepository:
    """
    Repository để làm việc với thông tin xử lý hàng loạt.
    """
    def __init__(self):
        self.batch_metadata_file = os.path.join(settings.TEMP_DIR, "excel_batch_processing_metadata.json")
        self.batches: Dict[str, BatchProcessingInfo] = {}
        self._load_metadata()

    def _load_metadata(self) -> None:
        try:
            if os.path.exists(self.batch_metadata_file):
                with open(self.batch_metadata_file, "r") as f:
                    data = json.load(f)
                    for batch_id, batch_data in data.items():
                        self.batches[batch_id] = BatchProcessingInfo(**batch_data)
        except Exception as e:
            print(f"Error loading Excel batch doc_metadata: {e}, creating new file if it doesn't exist.")
            if not os.path.exists(self.batch_metadata_file):
                self._save_metadata()

    def _save_metadata(self) -> None:
        try:
            os.makedirs(settings.TEMP_DIR, exist_ok=True)
            data = {batch_id: batch.dict() for batch_id, batch in self.batches.items()}
            with open(self.batch_metadata_file, "w") as f:
                json.dump(data, f, default=str)
        except Exception as e:
            raise StorageException(f"Không thể lưu doc_metadata xử lý hàng loạt Excel: {str(e)}")

    async def save(self, batch_info: BatchProcessingInfo) -> BatchProcessingInfo:
        try:
            if not batch_info.id:
                 batch_info.id = str(uuid.uuid4())
            if not batch_info.created_at:
                 batch_info.created_at = datetime.utcnow()
            self.batches[batch_info.id] = batch_info
            self._save_metadata()
            return batch_info
        except Exception as e:
            raise StorageException(f"Không thể lưu thông tin xử lý hàng loạt Excel: {str(e)}")

    async def get(self, batch_id: str) -> BatchProcessingInfo:
        try:
            if batch_id not in self.batches:
                raise DocumentNotFoundException(f"Batch processing info with id '{batch_id}' not found.") # Hoặc BatchNotFound
            return self.batches[batch_id]
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể lấy thông tin xử lý hàng loạt Excel {batch_id}: {str(e)}")

    async def update(self, batch_info: BatchProcessingInfo) -> BatchProcessingInfo:
        try:
            if batch_info.id not in self.batches:
                raise DocumentNotFoundException(f"Batch processing info with id '{batch_info.id}' not found for update.")

            self.batches[batch_info.id] = batch_info
            self._save_metadata()
            return batch_info
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể cập nhật thông tin xử lý hàng loạt Excel {batch_info.id}: {str(e)}")

    async def delete(self, batch_id: str) -> None:
        try:
            if batch_id not in self.batches:
                raise DocumentNotFoundException(f"Batch processing info with id '{batch_id}' not found for deletion.")
            del self.batches[batch_id]
            self._save_metadata()
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể xóa thông tin xử lý hàng loạt Excel {batch_id}: {str(e)}")

class MergeRepository:
    """
    Repository để làm việc với thông tin gộp tài liệu Excel (gộp sheet hoặc file).
    """
    def __init__(self):
        self.merge_metadata_file = os.path.join(settings.TEMP_DIR, "excel_merge_metadata.json")
        self.merges: Dict[str, MergeInfo] = {}
        self._load_metadata()

    def _load_metadata(self) -> None:
        try:
            if os.path.exists(self.merge_metadata_file):
                with open(self.merge_metadata_file, "r") as f:
                    data = json.load(f)
                    for merge_id, merge_data in data.items():
                        self.merges[merge_id] = MergeInfo(**merge_data)
        except Exception as e:
            print(f"Error loading Excel merge doc_metadata: {e}, creating new file if it doesn't exist.")
            if not os.path.exists(self.merge_metadata_file):
                self._save_metadata()

    def _save_metadata(self) -> None:
        try:
            os.makedirs(settings.TEMP_DIR, exist_ok=True)
            data = {merge_id: merge.dict() for merge_id, merge in self.merges.items()}
            with open(self.merge_metadata_file, "w") as f:
                json.dump(data, f, default=str)
        except Exception as e:
            raise StorageException(f"Không thể lưu doc_metadata gộp Excel: {str(e)}")

    async def save(self, merge_info: MergeInfo) -> MergeInfo:
        try:
            if not merge_info.id:
                 merge_info.id = str(uuid.uuid4())
            if not hasattr(merge_info, 'created_at') or not merge_info.created_at:
               
                pass
            self.merges[merge_info.id] = merge_info
            self._save_metadata()
            return merge_info
        except Exception as e:
            raise StorageException(f"Không thể lưu thông tin gộp Excel: {str(e)}")

    async def get(self, merge_id: str) -> MergeInfo:
        try:
            if merge_id not in self.merges:
                raise DocumentNotFoundException(f"Merge info with id '{merge_id}' not found.")
            return self.merges[merge_id]
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể lấy thông tin gộp Excel {merge_id}: {str(e)}")

    async def update(self, merge_info: MergeInfo) -> MergeInfo:
        try:
            if merge_info.id not in self.merges:
                raise DocumentNotFoundException(f"Merge info with id '{merge_info.id}' not found for update.")

            self.merges[merge_info.id] = merge_info
            self._save_metadata()
            return merge_info
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể cập nhật thông tin gộp Excel {merge_info.id}: {str(e)}")

    async def delete(self, merge_id: str) -> None:
        try:
            if merge_id not in self.merges:
                raise DocumentNotFoundException(f"Merge info with id '{merge_id}' not found for deletion.")
            del self.merges[merge_id]
            self._save_metadata()
        except DocumentNotFoundException:
            raise
        except Exception as e:
            raise StorageException(f"Không thể xóa thông tin gộp Excel {merge_id}: {str(e)}")